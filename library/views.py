
from django.views.generic import CreateView, ListView, View
from django.shortcuts import render
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.urls import reverse_lazy
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'form.html'
    success_url = reverse_lazy('author_list')

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library/form.html'
    success_url = '/books/'

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'form.html'
    success_url = reverse_lazy('borrow_list')

class AuthorListView(ListView):
    model = Author
    paginate_by = 5
    template_name = 'list.html'
    context_object_name = 'objects'

class BookListView(ListView):
    model = Book
    paginate_by = 5
    template_name = 'list.html'
    context_object_name = 'objects'

class BorrowListView(ListView):
    model = BorrowRecord
    paginate_by = 5
    template_name = 'borrowrecord_list.html'
    context_object_name = 'objects'

class ExportExcelView(View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Author Sheet
        authors = Author.objects.all()
        sheet = wb.create_sheet("Authors")
        sheet.append(['ID', 'Name', 'Email', 'Bio'])
        for a in authors:
            sheet.append([a.id, a.name, a.email, a.bio])

        # Book Sheet
        books = Book.objects.all()
        sheet = wb.create_sheet("Books")
        sheet.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
        for b in books:
            sheet.append([b.id, b.title, b.genre, b.published_date, b.author.name])

        # Borrow Sheet
        borrows = BorrowRecord.objects.all()
        sheet = wb.create_sheet("Borrow Records")
        sheet.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
        for br in borrows:
            sheet.append([br.id, br.user_name, br.book.title, br.borrow_date, br.return_date])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb.save(response)
        return response

def home(request):
    return render(request, 'home.html')
